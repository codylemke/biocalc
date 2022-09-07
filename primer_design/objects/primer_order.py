"""
Contains the PrimerOrder object.

Classes:
    PrimerOrder

-------------------------------------------------------------------------------
"""
__author__ = 'Cody Lemke'
__version__ = '2.0.0'

import re
from pathlib import Path
# Third Party Packages
import openpyxl  # https://openpyxl.readthedocs.io/en/stable/
import python_codon_tables as pct  # https://github.com/Edinburgh-Genome-Foundry/codon-usage-tables
# Local Modules
from primer_design.objects import (
    Variant,
    DNA,
    Primer,
    PCRReaction,
)
from primer_design.algorithms import (
    design_amplification_primers,
    design_sequencing_primers,
    design_overlapping_primers,
    design_end_to_end_primers,
    design_golden_mutagenesis_primers,
    design_gibson_primers,
    design_cpec_primers,
    design_golden_gate_primers,
    design_hifi_primers
)

BASE_DIR = Path(__file__).resolve().parent.parent.parent


# CLASSES ---------------------------------------------------------------------
class PrimerOrder:
    """Representation of a primer order"""
    
    def __init__(self, experiment_name):
        """Primer order constructor function"""
        self.configure_experiment(experiment_name)
        # self.experiment_dir
        # self.output_dir
        # self.input_template
        self.parse_input_template()
        # self.template_name
        # self.template
        # self.destination_taxon
        # self.polymerase
        # self.method
        # self.template_concentration
        # self.primer_concentration
        # self.reaction_volume
        # self.primer_supplier
        # self.primer_scale
        # self.min_primer_length
        # self.avoid_enzymes
        # self.circular_dna
        # self.restriction_enzyme
        # self.prefix_overhang
        # self.suffix
        # self.five_prime_gg_overhang
        # self.three_primer_gg_overhang
        # self.replacement_range
        # self.min_binding_length
        # self.max_binding_length
        # self.target_tm
        # self.construct_prefix
        # self.construct_starting_number
        # self.variant_codes
        self.codon_table = pct.get_codons_table(self.destination_taxon)
        self.generate_pcr_reactions()
        self.primers = self.experiment_dir / 'output' / 'primers.tsv'
        self.assembly = self.experiment_dir / 'output' / 'assembly.tsv'
        self.mutant_sequences = self.experiment_dir / 'output' / 'mutant_sequences.fasta'
        return

    def configure_experiment(self, experiment_name):
        """
        Configure the directory structure for the run and return the path to
        the input file
        """
        # Create a directory for experiments if it is not present
        EXPERIMENTS_DIR = BASE_DIR / 'experiments'
        if not EXPERIMENTS_DIR.is_dir():
            EXPERIMENTS_DIR.mkdir()
            print('Created "experiments/" directory')
        # Create a directory for the order if it is not present
        ORDER_DIR = EXPERIMENTS_DIR / experiment_name
        if not ORDER_DIR.is_dir():
            ORDER_DIR.mkdir()
            print(f'Created "experiments/{experiment_name}/" directory')
        # Create a directory for the output if it is not present
        OUTPUT_DIR = ORDER_DIR / 'output'
        if not OUTPUT_DIR.is_dir():
            OUTPUT_DIR.mkdir()
            print(f'Created "experiments/{experiment_name}/output" directory')
        # Select input template
        input_excel_file = None
        while input_excel_file is None:
            order_dir_excel_files = [file for file in ORDER_DIR.glob('*.xlsx')]
            if len(order_dir_excel_files) == 0:
                input("""
                    Please add your input template to the experiment directory
                    and press enter
                """)
            elif len(order_dir_excel_files) > 1:
                input('More than one excel file is present in your experiment directory\nRemove all excel files except your input file and press enter')
            else:
                input_excel_file = order_dir_excel_files[0]
        return input_excel_file

    def parse_input_template(self):
        """Returns a list of variant instances"""
        file = openpyxl.load_workbook(self.input_template, data_only=True)
        self.enzyme = 'enzyme'
        self.template = DNA(file['Inputs']['B3'].value, name=file['Inputs']['B2'].value) # template
        self.destination_taxon = file['Inputs']['B4'].value # target_organism
        self.polymerase = file['Inputs']['B5'].value # exclude
        self.purpose = 'mutagenesis'
        self.method = file['Inputs']['B6'].value # method
        self.template_concentration = file['Inputs']['B7'].value # exclude
        self.primer_concentration = file['Inputs']['B8'].value # exclude
        self.reaction_volume = file['Inputs']['B9'].value # exclude
        self.primer_supplier = file['Inputs']['B10'].value # supplier
        self.primer_scale = file['Inputs']['B11'].value # primer_scale
        self.min_primer_length = file['Inputs']['B13'].value # constraint
        self.avoid_enzymes = file['Inputs']['B14'].value # constraint
        self.circular_dna = file['Inputs']['B15'].value # determined via DNA type
        self.restriction_enzyme = file['Inputs']['B17'].value 
        self.prefix_overhang = file['Inputs']['B18'].value # algorithmically determined
        self.suffix = file['Inputs']['B19'].value # algorithmically determined
        self.five_prime_gg_overhang = file['Inputs']['B20'].value # ?
        self.three_primer_gg_overhang = file['Inputs']['B21'].value # ?
        self.replacement_range = file['Inputs']['B22'].value
        self.min_binding_length = file['Inputs']['B23'].value
        self.max_binding_length = file['Inputs']['B24'].value
        self.target_tm = file['Inputs']['B25'].value
        self.construct_prefix = file['Mutations']['B1'].value # name
        self.construct_starting_number = file['Mutations']['B2'].value # primary_key
        self.variant_codes = parse_variant_input(file)


        def parse_variant_input(file):
            """Return a list of variant codes included in the complex variant code"""
            # Parse mutations input
            complex_variant_codes = list()
            for row in range(4, file['Mutations'].max_row+1):
                complex_variant_code = file['Mutations'].cell(row=row, column=1).value.strip()
                complex_variant_codes.append(complex_variant_code)
            variant_codes = list()
            # Deconstruct the complex variant codes
            for complex_variant_code in complex_variant_codes:
                aa_position = re.search(r'\d+', complex_variant_code).group()
                wt_residue = complex_variant_code.split(aa_position)[0]
                variant_residues = complex_variant_code.split(aa_position)[1]
                if '(' in variant_residues and ')' in variant_residues:
                    # Process degenerate codons
                    degenerate_codons = variant_residues.replace('(', '').split(')')[:-1]
                    for degenerate_codon in degenerate_codons:
                        variant_code = f'{wt_residue}{aa_position}({degenerate_codon})'
                        variant_codes.append(variant_code)
                else:
                    # Proceed normally
                    for variant_residue in variant_residues:
                        variant_code = f'{wt_residue}{aa_position}{variant_residue}'
                        variant_codes.append(variant_code)
            return variant_codes

        return None

    def generate_pcr_reactions(self):
        """Return list of pcr reactions from the order"""
        pcr_reactions = list()
        for variant_code in self.variant_codes:
            aa_position = re.search(r'\d+', variant_code).group()
            wt_residue = variant_code.split(aa_position)[0]
            variation = variant_code.split(aa_position)[1]
            codon_index = (int(aa_position)*3 - 3, int(aa_position)*3)
            # Trick Primers
            if variation == '#':
                reaction_type = 'trick'
                primer_codons = ['NDT','VMA','ATG','TGG']
                possible_codons = list()
                possible_residues = list()
                for codon in primer_codons:
                    possible_codons.extend(parse_degenerate_codon(codon))
                # Create Variants
                for possible_codon in possible_codons:
                    translated_codon = translate_codon(possible_codon)
                    if translated_codon not in possible_residues:
                        possible_residues.append(translated_codon)
                possible_variants = list()
                for possible_residue in possible_residues:
                    possible_variant = Variant(self.enzyme, f'{wt_residue}{aa_position}{possible_residue}')
                    possible_variants.append(possible_variant)
                # Create Primers
                primer_pairs = list()
                for primer_codon in primer_codons:
                    primer_pair = self.design_primers(codon_index=codon_index, codon=primer_codon)
                    primer_pairs.append(primer_pair)
                primers = list()
                for primer_pair in primer_pairs:
                    for primer in primer_pair:
                        primers.append(primer)
            # Smart Primers
            elif variation == '!':
                reaction_type = 'smart'
                primer_codons = ['NDT','VHG','TGG']
                possible_codons = list()
                possible_residues = list()
                for codon in primer_codons:
                    possible_codons.extend(parse_degenerate_codon(codon))
                # Create Variants
                for possible_codon in possible_codons:
                    translated_codon = translate_codon(possible_codon)
                    if translated_codon not in possible_residues:
                        possible_residues.append(translated_codon)
                variants = list()
                for possible_residue in possible_residues:
                    variant = Variant(self.enzyme, f'{wt_residue}{aa_position}{possible_residue}')
                    variants.append(variant)
                # Create Primers
                primer_pairs = list()
                for primer_codon in primer_codons:
                    primer_pair = self.design_primers(codon_index=codon_index, codon=primer_codon)
                    primer_pairs.append(primer_pair)
                primers = list()
                for primer_pair in primer_pairs:
                    for primer in primer_pair:
                        primers.append(primer)
            # Degenerate Primers
            elif len(variation) > 1:
                reaction_type = 'degenerate'
                primer_codon = variation.replace('(', '').replace(')', '')
                possible_codons = parse_degenerate_codon(primer_codon)
                # Create Variants
                possible_residues = list()
                for possible_codon in possible_codons:
                    translated_codon = translate_codon(possible_codon)
                    if translated_codon not in possible_residues:
                        possible_residues.append(translated_codon)
                variants = list()
                for possible_residue in possible_residues:
                    variant = Variant(self.enzyme, f'{wt_residue}{aa_position}{possible_residue}')
                    variants.append(variant)
                # Create Primers
                primer_pair = self.design_primers(codon_index=codon_index, codon=primer_codon)
                primers = list()
                for primer in primer_pair:
                    primers.append(primer)
            # Standard Primers
            else:
                reaction_type = 'standard'
                variant = Variant(self.enzyme, variant_code)
                # Determine best codon
                best_codon_frequency = 0
                for codon in self.codon_table[variation].keys():
                    codon_frequency = self.codon_table[variation][codon]
                    if codon_frequency > best_codon_frequency:
                        best_codon = codon
                        codon_frequency = codon_frequency
                # Create Primers
                primer_pair = self.design_primers(codon_index=codon_index, codon=best_codon)
                primers = list()
                for primer in primer_pair:
                    primers.append(primer)
            # Create PCR Reaction Object
            pcr_reaction = PCRReaction(
                reaction_volume=self.reaction_volume,
                polymerase=self.polymerase,
                template=self.template,
                template_concentration=self.template_concentration,
                primers=primers,
                reaction_type=reaction_type,
            )
            pcr_reactions.append(pcr_reaction)


        def parse_degenerate_codon(degenerate_codon):
            """Return a list of codons resulting from a degenerate codon"""
            degenerate_codon_table = {
                'R': ['A', 'G'],
                'Y': ['C', 'T'],
                'M': ['A', 'C'],
                'K': ['G', 'T'],
                'S': ['C', 'G'],
                'W': ['A', 'T'],
                'H': ['A', 'C', 'T'],
                'B': ['C', 'G', 'T'],
                'V': ['A', 'C', 'G'],
                'D': ['A', 'G', 'T'],
                'N': ['A', 'C', 'G', 'T'],
            }
            position_one = degenerate_codon_table[degenerate_codon[0]]
            position_two = degenerate_codon_table[degenerate_codon[1]]
            position_three = degenerate_codon_table[degenerate_codon[2]]
            codons = list()
            for nucleotide_one in position_one:
                for nucleotide_two in position_two:
                    for nucleotide_three in position_three:
                        codon = f'{nucleotide_one}{nucleotide_two}{nucleotide_three}'
                        codons.append(codon)
            return codons

        def translate_codon(codon):
            """Return the amino acid residue a codon codes for in a translation_table"""
            for residue in self.codon_table.keys():
                if codon in self.codon_table[residue].keys():
                    return residue

        return pcr_reactions

    def design_primers(self, codon_index=None, codon=None):
        """Calls the appropriate primer design method based on primer purpose"""
        # Constraints
        self.FIVE_PRIME_EDGE_GC_CLAMP = False
        self.THREE_PRIME_EDGE_GC_CLAMP = True
        self.MIN_PRIMER_LENGTH = 18
        self.MAX_PRIMER_LENGTH = 60
        self.MIN_PRIMER_TM = 53
        self.MAX_PRIMER_TM = 70
        self.MIN_PRIMER_TM2 = 58
        self.MAX_PRIMER_TM2 = 95
        self.MIN_PRIMER_GC_CONTENT = 30
        self.MAX_PRIMER_GC_CONTENT = 70
        self.MIN_FIVE_PRIME_EDGE_LENGTH = 11
        self.MAX_FIVE_PRIME_EDGE_LENGTH = 40
        self.MIN_THREE_PRIME_EDGE_LENGTH = 11
        self.MAX_THREE_PRIME_EDGE_LENGTH = 40
        self.MIN_FIVE_PRIME_EDGE_TM = 35
        self.MAX_FIVE_PRIME_EDGE_TM = 55
        self.MIN_THREE_PRIME_EDGE_TM = 35
        self.MAX_THREE_PRIME_EDGE_TM = 55
        # Decision Tree
        if self.purpose == 'amplification':
            if self.method == 'amplification':
                # Parameters
                self.OPTIMAL_PRIMER_LENGTH = 36
                self.OPTIMAL_PRIMER_TM = 60
                self.OPTIMAL_PRIMER_TM2 = 70
                self.OPTIMAL_PRIMER_GC = 50
                # Scoring Weight
                self.PRIMER_LENGTH_BIAS = 1
                self.PRIMER_TM_BIAS = 1
                self.PRIMER_TM2_BIAS = 1
                self.PRIMER_GC_BIAS = 1
                primers = design_amplification_primers(self)
            else:
                raise ValueError('Invalid amplification design method')
        elif self.purpose == 'sequencing':
            if self.method == 'sequencing':
                # Set Optimal Parameters
                self.OPTIMAL_PRIMER_LENGTH = 36
                self.OPTIMAL_PRIMER_TM = 60
                self.OPTIMAL_PRIMER_TM2 = 70
                self.OPTIMAL_PRIMER_GC = 50
                # Scoring Weight
                self.PRIMER_LENGTH_BIAS = 1
                self.PRIMER_TM_BIAS = 1
                self.PRIMER_TM2_BIAS = 1
                self.PRIMER_GC_BIAS = 1
                primers = design_sequencing_primers(self)
            else:
                raise ValueError('Invalid sequencing design method')
        elif self.purpose == 'mutagenesis':
            if self.method == 'overlapping':
                # Parameters
                self.OPTIMAL_PRIMER_LENGTH = 36
                self.OPTIMAL_PRIMER_TM = 60
                self.OPTIMAL_PRIMER_TM2 = 70
                self.OPTIMAL_PRIMER_GC = 50
                self.OPTIMAL_FIVE_PRIME_EDGE_LENGTH = 21
                self.OPTIMAL_THREE_PRIME_EDGE_LENGTH = 15
                self.OPTIMAL_FIVE_PRIME_EDGE_TM = 50
                self.OPTIMAL_THREE_PRIME_EDGE_TM = 42
                # Scoring Weight
                self.PRIMER_LENGTH_BIAS = 1
                self.PRIMER_TM_BIAS = 1
                self.PRIMER_TM2_BIAS = 1
                self.PRIMER_GC_BIAS = 1
                self.FIVE_PRIME_EDGE_LENGTH_BIAS = 1
                self.THREE_PRIME_EDGE_LENGTH_BIAS = 1
                self.FIVE_PRIME_EDGE_TM_BIAS = 1
                self.THREE_PRIME_EDGE_TM_BIAS = 1
                primers = design_overlapping_primers(self, codon_index, codon)
            elif self.method == 'end_to_end':
                # Parameters
                self.OPTIMAL_PRIMER_LENGTH = 36
                self.OPTIMAL_PRIMER_TM = 60
                self.OPTIMAL_PRIMER_TM2 = 70
                self.OPTIMAL_PRIMER_GC = 50
                self.OPTIMAL_FIVE_PRIME_EDGE_LENGTH = 21
                self.OPTIMAL_THREE_PRIME_EDGE_LENGTH = 15
                self.OPTIMAL_FIVE_PRIME_EDGE_TM = 50
                self.OPTIMAL_THREE_PRIME_EDGE_TM = 42
                # Scoring Weight
                self.PRIMER_LENGTH_BIAS = 1
                self.PRIMER_TM_BIAS = 1
                self.PRIMER_TM2_BIAS = 1
                self.PRIMER_GC_BIAS = 1
                self.FIVE_PRIME_EDGE_LENGTH_BIAS = 1
                self.THREE_PRIME_EDGE_LENGTH_BIAS = 1
                self.FIVE_PRIME_EDGE_TM_BIAS = 1
                self.THREE_PRIME_EDGE_TM_BIAS = 1
                primers = design_end_to_end_primers(self, codon_index, codon)
            elif self.method == 'golden_mutagenesis':
                # Parameters
                self.OPTIMAL_PRIMER_LENGTH = 36
                self.OPTIMAL_PRIMER_TM = 60
                self.OPTIMAL_PRIMER_TM2 = 70
                self.OPTIMAL_PRIMER_GC = 50
                self.OPTIMAL_FIVE_PRIME_EDGE_LENGTH = 21
                self.OPTIMAL_THREE_PRIME_EDGE_LENGTH = 15
                self.OPTIMAL_FIVE_PRIME_EDGE_TM = 50
                self.OPTIMAL_THREE_PRIME_EDGE_TM = 42
                # Scoring Weight
                self.PRIMER_LENGTH_BIAS = 1
                self.PRIMER_TM_BIAS = 1
                self.PRIMER_TM2_BIAS = 1
                self.PRIMER_GC_BIAS = 1
                self.FIVE_PRIME_EDGE_LENGTH_BIAS = 1
                self.THREE_PRIME_EDGE_LENGTH_BIAS = 1
                self.FIVE_PRIME_EDGE_TM_BIAS = 1
                self.THREE_PRIME_EDGE_TM_BIAS = 1
                primers = design_golden_mutagenesis_primers(self, codon_index, codon)
            else:
                raise ValueError('Invalid mutagenesis design method')
        elif self.purpose == 'assembly':
            if self.method == 'gibson':
                # Parameters
                self.OPTIMAL_PRIMER_LENGTH = 36
                self.OPTIMAL_PRIMER_TM = 60
                self.OPTIMAL_PRIMER_TM2 = 70
                self.OPTIMAL_PRIMER_GC = 50
                self.OPTIMAL_FIVE_PRIME_EDGE_LENGTH = 21
                self.OPTIMAL_THREE_PRIME_EDGE_LENGTH = 15
                self.OPTIMAL_FIVE_PRIME_EDGE_TM = 50
                self.OPTIMAL_THREE_PRIME_EDGE_TM = 42
                # Scoring Weight
                self.PRIMER_LENGTH_BIAS = 1
                self.PRIMER_TM_BIAS = 1
                self.PRIMER_TM2_BIAS = 1
                self.PRIMER_GC_BIAS = 1
                self.FIVE_PRIME_EDGE_LENGTH_BIAS = 1
                self.THREE_PRIME_EDGE_LENGTH_BIAS = 1
                self.FIVE_PRIME_EDGE_TM_BIAS = 1
                self.THREE_PRIME_EDGE_TM_BIAS = 1
                primers = design_gibson_primers(self)
            elif self.method == 'cpec':
                # Parameters
                self.OPTIMAL_PRIMER_LENGTH = 36
                self.OPTIMAL_PRIMER_TM = 60
                self.OPTIMAL_PRIMER_TM2 = 70
                self.OPTIMAL_PRIMER_GC = 50
                self.OPTIMAL_FIVE_PRIME_EDGE_LENGTH = 21
                self.OPTIMAL_THREE_PRIME_EDGE_LENGTH = 15
                self.OPTIMAL_FIVE_PRIME_EDGE_TM = 50
                self.OPTIMAL_THREE_PRIME_EDGE_TM = 42
                # Scoring Weight
                self.PRIMER_LENGTH_BIAS = 1
                self.PRIMER_TM_BIAS = 1
                self.PRIMER_TM2_BIAS = 1
                self.PRIMER_GC_BIAS = 1
                self.FIVE_PRIME_EDGE_LENGTH_BIAS = 1
                self.THREE_PRIME_EDGE_LENGTH_BIAS = 1
                self.FIVE_PRIME_EDGE_TM_BIAS = 1
                self.THREE_PRIME_EDGE_TM_BIAS = 1
                primers = design_cpec_primers(self)
            elif self.method == 'golden_gate':
                # Parameters
                self.OPTIMAL_PRIMER_LENGTH = 36
                self.OPTIMAL_PRIMER_TM = 60
                self.OPTIMAL_PRIMER_TM2 = 70
                self.OPTIMAL_PRIMER_GC = 50
                self.OPTIMAL_FIVE_PRIME_EDGE_LENGTH = 21
                self.OPTIMAL_THREE_PRIME_EDGE_LENGTH = 15
                self.OPTIMAL_FIVE_PRIME_EDGE_TM = 50
                self.OPTIMAL_THREE_PRIME_EDGE_TM = 42
                # Scoring Weight
                self.PRIMER_LENGTH_BIAS = 1
                self.PRIMER_TM_BIAS = 1
                self.PRIMER_TM2_BIAS = 1
                self.PRIMER_GC_BIAS = 1
                self.FIVE_PRIME_EDGE_LENGTH_BIAS = 1
                self.THREE_PRIME_EDGE_LENGTH_BIAS = 1
                self.FIVE_PRIME_EDGE_TM_BIAS = 1
                self.THREE_PRIME_EDGE_TM_BIAS = 1
                primers = design_golden_gate_primers(self)
            elif self.method == 'hifi':
                # Parameters
                self.OPTIMAL_PRIMER_LENGTH = 36
                self.OPTIMAL_PRIMER_TM = 60
                self.OPTIMAL_PRIMER_TM2 = 70
                self.OPTIMAL_PRIMER_GC = 50
                self.OPTIMAL_FIVE_PRIME_EDGE_LENGTH = 21
                self.OPTIMAL_THREE_PRIME_EDGE_LENGTH = 15
                self.OPTIMAL_FIVE_PRIME_EDGE_TM = 50
                self.OPTIMAL_THREE_PRIME_EDGE_TM = 42
                # Scoring Weight
                self.PRIMER_LENGTH_BIAS = 1
                self.PRIMER_TM_BIAS = 1
                self.PRIMER_TM2_BIAS = 1
                self.PRIMER_GC_BIAS = 1
                self.FIVE_PRIME_EDGE_LENGTH_BIAS = 1
                self.THREE_PRIME_EDGE_LENGTH_BIAS = 1
                self.FIVE_PRIME_EDGE_TM_BIAS = 1
                self.THREE_PRIME_EDGE_TM_BIAS = 1
                primers = design_hifi_primers(self)
            else:
                raise ValueError('Invalid assembly design method')
        else:
            raise ValueError('primer purpose is invalid')
        
        return primers

    def validate_primer(self, primer):
        """Returns boolean indicating whether the input primer is within all the primer constraints"""
        if primer.length < self.MIN_PRIMER_LENGTH:
            return False
        elif primer.length > self.MAX_PRIMER_LENGTH:
            return False
        elif primer.tm < self.MIN_PRIMER_TM:
            return False
        elif primer.tm > self.MAX_PRIMER_TM:
            return False
        elif primer.tm2 < self.MIN_PRIMER_TM2:
            return False
        elif primer.tm2 > self.MAX_PRIMER_TM2:
            return False
        elif self.FIVE_PRIME_EDGE_GC_CLAMP and (primer.sequence[0] == 'A' or primer.sequence[0] == 'T'):
            return False
        elif self.THREE_PRIME_EDGE_GC_CLAMP and (primer.sequence[-1] == 'A' or primer.sequence[-1] == 'T'):
            return False
        elif primer.gc_content < self.MIN_PRIMER_GC_CONTENT:
            return False
        elif primer.gc_content > self.MAX_PRIMER_GC_CONTENT:
            return False        
        elif primer.five_prime_edge_length < self.MIN_FIVE_PRIME_EDGE_LENGTH:
            return False
        elif primer.five_prime_edge_length > self.MAX_FIVE_PRIME_EDGE_LENGTH:
            return False
        elif primer.three_prime_edge_length < self.MIN_THREE_PRIME_EDGE_LENGTH:
            return False
        elif primer.three_prime_edge_length > self.MAX_THREE_PRIME_EDGE_LENGTH:
            return False
        elif primer.five_prime_edge_tm < self.MIN_FIVE_PRIME_EDGE_TM:
            return False
        elif primer.five_prime_edge_tm > self.MAX_FIVE_PRIME_EDGE_TM:
            return False
        elif primer.three_prime_edge_tm < self.MIN_THREE_PRIME_EDGE_TM:
            return False
        elif primer.three_prime_edge_tm > self.MAX_THREE_PRIME_EDGE_TM:
            return False
        else:
            return True

    def amplify(pcr_reaction):
        """Placeholder"""
        fragments = list
        # do stuff
        return fragments

    def assemble(fragment, *args):
        """Placeholder"""
        construct = None
        return construct

    def transform(construct, strain):
        """Placeholder"""
        new_strain = None
        return new_strain

    def inoculate():
        pass

    def miniprep():
        pass

