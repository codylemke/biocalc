import openpyxl
from openpyxl import load_workbook
import collections
import math
import os
from pathlib import Path


def output_nanohive_excel(paths, assembly, volume, f, construct):
    """
    volume = 20 is the default reaction volume
    f = 0.05 # Final primer concentration
    construct = 5 # Final construct concentration

    creates assembly list
    """

    template_file_path = Path(__file__).resolve().parent.parent / 'Template Files' / 'Assembly Template.xlsx'
    wb = load_workbook(filename=template_file_path)
    ws = wb['Assembly List']

    ordered_assembly = collections.OrderedDict(sorted(assembly.items()))
    units = ['Mass (ng)', 'Volume (µL)', 'Concentration (µM)', 'Concentration (ng/µL)',
             'Moles (nmols)', 'Moles (µmols)', 'Moles (pmols)']

    single_conc = [construct, f, f]
    conc_smart = [[construct, f * 12 / 20, f * 12 / 20], [f * 6 / 20, f * 6 / 20], [f * 1 / 20, f * 1 / 20],
                  [f * 1 / 20, f * 1 / 20]]
    conc_trick = [[construct, f * 12 / 22, f * 12 / 22], [f * 9 / 22, f * 9 / 22], [f * 1 / 22, f * 1 / 22]]
    single_units = [units[3], units[2], units[2]]
    smart_units = [[units[3], units[2], units[2]], [units[2], units[2]], [units[2], units[2]], [units[2], units[2]]]
    trick_units = [[units[3], units[2], units[2]], [units[2], units[2]], [units[2], units[2]]]

    smart_count = 0
    trick_count = 0
    c = 2
    for key, values in ordered_assembly.items():
        if values[6] in ['Single', 'Degenerate']:
            for i in range(3):
                ws.cell(row=c + i, column=4).value = key
                ws.cell(row=c + i, column=5).value = volume
                ws.cell(row=c + i, column=6).value = values[i]
                ws.cell(row=c + i, column=7).value = single_conc[i]
                ws.cell(row=c + i, column=8).value = single_units[i]
            c = c + 3
        if values[6] == 'SMART':
            if smart_count == 0:
                for i in range(
                        3):  # [seq_name, primer_fname, primer_rname, mutant, str(primer_dict['tm']), str(primer_dict['tm_target']), primer_type_temp]
                    ws.cell(row=c + i, column=4).value = key[:-1]
                    ws.cell(row=c + i, column=5).value = volume
                    ws.cell(row=c + i, column=6).value = values[i]
                    ws.cell(row=c + i, column=7).value = conc_smart[smart_count][i]
                    ws.cell(row=c + i, column=8).value = smart_units[smart_count][i]
                c = c + 3
            if smart_count >= 1 and smart_count <= 3:
                for i in range(
                        2):  # [seq_name, primer_fname, primer_rname, mutant, str(primer_dict['tm']), str(primer_dict['tm_target']), primer_type_temp]
                    ws.cell(row=c + i, column=4).value = key[:-1]
                    ws.cell(row=c + i, column=5).value = volume
                    ws.cell(row=c + i, column=6).value = values[i + 1]
                    ws.cell(row=c + i, column=7).value = conc_smart[smart_count][i]
                    ws.cell(row=c + i, column=8).value = smart_units[smart_count][i]
                c = c + 2
            # Add to smart_count each loop
            smart_count += 1
            # On the fourth loop set count to zero because the SMART set is complete (4 degenerate primer sets)
            if smart_count == 4:
                smart_count = 0

        if values[6] == 'TRICK':
            if trick_count == 0:
                for i in range(
                        3):  # [seq_name, primer_fname, primer_rname, mutant, str(primer_dict['tm']), str(primer_dict['tm_target']), primer_type_temp]
                    ws.cell(row=c + i, column=4).value = key[:-1]
                    ws.cell(row=c + i, column=5).value = volume
                    ws.cell(row=c + i, column=6).value = values[i]
                    ws.cell(row=c + i, column=7).value = conc_trick[trick_count][i]
                    ws.cell(row=c + i, column=8).value = trick_units[trick_count][i]
                c = c + 3
            if trick_count >= 1 and trick_count <= 2:
                for i in range(
                        2):  # [seq_name, primer_fname, primer_rname, mutant, str(primer_dict['tm']), str(primer_dict['tm_target']), primer_type_temp]
                    ws.cell(row=c + i, column=4).value = key[:-1]
                    ws.cell(row=c + i, column=5).value = volume
                    ws.cell(row=c + i, column=6).value = values[i + 1]
                    ws.cell(row=c + i, column=7).value = conc_trick[trick_count][i]
                    ws.cell(row=c + i, column=8).value = trick_units[trick_count][i]
                c = c + 2
            # Add to trick_count each loop
            trick_count += 1
            # On the fourth loop set count to zero because the TRICK set is complete (3 degenerate primer sets)
            if trick_count == 3:
                trick_count = 0

    ws.cell(row=c, column=4).value = 'Control'
    ws.cell(row=c, column=5).value = volume
    ws.cell(row=c, column=6).value = 'Control_construct'
    ws.cell(row=c, column=7).value = construct
    ws.cell(row=c, column=8).value = 'Concentration (ng/µL)'

    ws.cell(row=c+1, column=4).value = 'Control'
    ws.cell(row=c+1, column=5).value = volume
    ws.cell(row=c+1, column=6).value = 'Control_F'
    ws.cell(row=c+1, column=7).value = f
    ws.cell(row=c+1, column=8).value = 'Concentration (µM)'

    ws.cell(row=c+2, column=4).value = 'Control'
    ws.cell(row=c+2, column=5).value = volume
    ws.cell(row=c+2, column=6).value = 'Control_R'
    ws.cell(row=c+2, column=7).value = f
    ws.cell(row=c+2, column=8).value = 'Concentration (µM)'

    wb.save(paths['output'] / 'Nanohive_Assembly.xlsx')


def output_nanohive_excel2(paths, assembly, volume, f, construct, stock_volume=200):
    """
    creates all the other 
    """
    template_file_path = paths['output'] / 'Nanohive_Assembly.xlsx'
    wb = load_workbook(filename=template_file_path)
    ordered_assembly = collections.OrderedDict(sorted(assembly.items()))

    mole_ratio_smart = [(12 / 20), (6 / 20), (1 / 20), (1 / 20)]
    mole_ratio_trick = [(12 / 22), (9 / 22), (1 / 22)]
    smart_volumes = [.5*(volume/20), .3*(volume/20), .1*(volume/20), .1*(volume/20)]
    trick_volumes = [.545454*(volume/20), .454545*(volume/20), .11363635*(volume/20)]

    for forward_or_reverse in [1, 2]:
        smart_count = 0
        trick_count = 0
        c = 3
        plate_count = forward_or_reverse
        for key, values in ordered_assembly.items():
            ws = wb['Plate ' + str(plate_count) + ' <' + str(plate_count) + '>']
            if values[6] in ['Single', 'Degenerate']:
                # Add to a list based on counter
                ws.cell(row=c, column=2).value = values[forward_or_reverse]
                ws.cell(row=c, column=5).value = f / 0.05
                ws.cell(row=c, column=6).value = 'Concentration (µM)'
                ws.cell(row=c, column=7).value = stock_volume
                c += 1
            if values[6] == 'SMART':
                ws.cell(row=c, column=2).value = values[forward_or_reverse]
                ws.cell(row=c, column=5).value = mole_ratio_smart[smart_count] * f * (volume) / smart_volumes[
                    smart_count]
                ws.cell(row=c, column=6).value = 'Concentration (µM)'
                ws.cell(row=c, column=7).value = stock_volume
                c += 1
                smart_count += 1
                # On the fourth loop set count to zero because the SMART set is complete (4 degenerate primer sets)
                if smart_count == 4:
                    smart_count = 0
            if values[6] == 'TRICK':
                ws.cell(row=c, column=2).value = values[forward_or_reverse]
                ws.cell(row=c, column=5).value = mole_ratio_trick[trick_count] * f * (volume) / trick_volumes[
                    trick_count]
                ws.cell(row=c, column=6).value = 'Concentration (µM)'
                ws.cell(row=c, column=7).value = stock_volume
                c += 1
                trick_count += 1
                # On the fourth loop set count to zero because the TRICK set is complete (3 degenerate primer sets)
                if trick_count == 3:
                    trick_count = 0
            # if (plate_count == 7 or plate_count == 8) and c - 3 == 96:
            #     print("WARNING: Too many primers; output files likely have errors")
            if c - 3 == 96:
                plate_count += 2
                c = 3

    # Add vectors and controls to plate zero
    ws = wb['Plate C <C>']
    ws.cell(row=3, column=2).value = 'Control_F'
    ws.cell(row=3, column=5).value = f / 0.05
    ws.cell(row=3, column=6).value = 'Concentration (µM)'
    ws.cell(row=3, column=7).value = stock_volume

    ws.cell(row=4, column=2).value = 'Control_R'
    ws.cell(row=4, column=5).value = f / 0.05
    ws.cell(row=4, column=6).value = 'Concentration (µM)'
    ws.cell(row=4, column=7).value = stock_volume

    ws.cell(row=5, column=2).value = 'Control_construct'
    ws.cell(row=5, column=5).value = construct * 20
    ws.cell(row=5, column=6).value = 'Concentration (ng/µL)'
    ws.cell(row=5, column=7).value = stock_volume - 20

    ws = wb['Plate 0 <0>']
    construct_well_count = int(math.ceil(len(ordered_assembly)/(stock_volume - 20)))  # safety factor for 10 uL dead volume in nanohive

    for c in range(3, 3+construct_well_count):
        ws.cell(row=c, column=2).value = list(ordered_assembly.items())[0][1][0]  # Get construct name
        ws.cell(row=c, column=5).value = construct * 20
        ws.cell(row=c, column=6).value = 'Concentration (ng/µL)'
        ws.cell(row=c, column=7).value = stock_volume - 20

    wb.save(paths['output'] / 'Nanohive_Assembly.xlsx')


def output_biomek_excel_files(paths, assembly, volume, f):
    ordered_assembly = collections.OrderedDict(sorted(assembly.items()))
    mole_ratio_smart = [(12 / 20), (6 / 20), (1 / 20), (1 / 20)]
    mole_ratio_trick = [(12 / 22), (9 / 22), (1 / 22)]
    smart_volumes = [.5*(volume/20), .3*(volume/20), .1*(volume/20), .1*(volume/20)]
    trick_volumes = [.545454*(volume/20), .454545*(volume/20), .11363635*(volume/20)]
    well_nums = ['ABCDEFGH'[idx // 12] + str(idx % 12 + 1) for idx in range(96)]

    for forward_or_reverse in [1, 2]:
        smart_count = 0
        trick_count = 0
        c = 2
        plate_count = forward_or_reverse
        wb = openpyxl.Workbook()
        ws = wb.create_sheet()
        ws.title = f"Plate {plate_count}"
        add_names_to_excel_sheet(ws)
        for key, values in ordered_assembly.items():
            if values[6] in ['Single', 'Degenerate']:
                # Add to a list based on counter
                ws.cell(row=c, column=7).value = values[forward_or_reverse]
                ws.cell(row=c, column=1).value = plate_count
                ws.cell(row=c, column=2).value = well_nums[int((c - 2))]
                ws.cell(row=c, column=3).value = (f / 0.05)*100/5
                ws.cell(row=c, column=4).value = 'Intermediate_Plate'
                ws.cell(row=c, column=5).value = well_nums[int((c - 2))]
                ws.cell(row=c, column=6).value = (f / 0.05)
                c += 1
            if values[6] == 'SMART':
                ws.cell(row=c, column=7).value = values[forward_or_reverse]
                ws.cell(row=c, column=1).value = plate_count
                ws.cell(row=c, column=2).value = well_nums[int((c - 2))]
                con_temp = (mole_ratio_smart[smart_count] * f * (volume) / smart_volumes[smart_count])
                ws.cell(row=c, column=3).value = con_temp * 100 / 5
                ws.cell(row=c, column=4).value = 'Intermediate_Plate'
                ws.cell(row=c, column=5).value = well_nums[int((c - 2))]
                ws.cell(row=c, column=6).value = con_temp
                c += 1
                smart_count += 1
                # On the fourth loop set count to zero because the SMART set is complete (4 degenerate primer sets)
                if smart_count == 4:
                    smart_count = 0
            if values[6] == 'TRICK':
                ws.cell(row=c, column=7).value = values[forward_or_reverse]
                ws.cell(row=c, column=1).value = plate_count
                ws.cell(row=c, column=2).value = well_nums[int((c - 2))]
                con_temp = (mole_ratio_trick[trick_count] * f * (volume) / trick_volumes[trick_count])
                ws.cell(row=c, column=3).value = con_temp * 100 / 5
                ws.cell(row=c, column=4).value = 'Intermediate_Plate'
                ws.cell(row=c, column=5).value = well_nums[int((c - 2))]
                ws.cell(row=c, column=6).value = con_temp
                c += 1
                trick_count += 1
                # On the fourth loop set count to zero because the TRICK set is complete (3 degenerate primer sets)
                if trick_count == 3:
                    trick_count = 0
            if well_nums[int(c - 2)-1] == 'H12':
                wb.remove(wb['Sheet'])
                wb.save(paths['output'] / f'Biomek_transfers_primers_plate_{plate_count}.xlsx')
                save_data_to_csv(paths['output'] / f'Biomek_transfers_primers_plate_{plate_count}.xlsx',
                                 paths['output'] / f'Biomek_transfers_primers_plate_{plate_count}.csv')
                plate_count += 2
                wb = openpyxl.Workbook()
                ws = wb.create_sheet()
                ws.title = f"Plate {plate_count}"
                add_names_to_excel_sheet(ws)
                c = 2
            # if (plate_count == 7 or plate_count == 8) and c == 98:
            #     print("WARNING: Too many primers; output files likely have errors")
        wb.remove(wb['Sheet'])
        wb.save(paths['output'] / f'Biomek_transfers_primers_plate_{plate_count}.xlsx')
        save_data_to_csv(paths['output'] / f'Biomek_transfers_primers_plate_{plate_count}.xlsx',
                         paths['output'] / f'Biomek_transfers_primers_plate_{plate_count}.csv')

    for forward_or_reverse in [1, 2]:
        smart_count = 0
        trick_count = 0
        c = 2
        plate_count = forward_or_reverse
        wb = openpyxl.Workbook()
        ws = wb.create_sheet()
        ws.title = f"Plate {plate_count}"
        ws.cell(row=1, column=1).value = 'Family ID'
        ws.cell(row=1, column=2).value = 'Well ID'
        ws.cell(row=1, column=3).value = 'Volume'
        ws.cell(row=1, column=4).value = 'Reservoir Number'
        ws.cell(row=1, column=5).value = 'Reservoir Well'
        ws.cell(row=1, column=6).value = 'Concentration'
        ws.cell(row=1, column=7).value = 'Primer ID'
        for key, values in ordered_assembly.items():
            if values[6] in ['Single', 'Degenerate']:
                # Add to a list based on counter
                ws.cell(row=c, column=1).value = plate_count
                ws.cell(row=c, column=2).value = well_nums[int(c - 2)]
                ws.cell(row=c, column=3).value = 100 - ((f / 0.05) * 100 / 5)
                ws.cell(row=c, column=4).value = 'water'
                ws.cell(row=c, column=5).value = 1
                ws.cell(row=c, column=6).value = (f / 0.05)
                ws.cell(row=c, column=7).value = values[forward_or_reverse]
                c += 1
            if values[6] == 'SMART':
                ws.cell(row=c, column=7).value = values[forward_or_reverse]
                ws.cell(row=c, column=1).value = plate_count
                ws.cell(row=c, column=2).value = well_nums[int(c - 2)]
                con_temp = (mole_ratio_smart[smart_count] * f * (volume) / smart_volumes[smart_count])
                ws.cell(row=c, column=3).value = 100 - (con_temp * 100 / 5)
                ws.cell(row=c, column=4).value = 'water'
                ws.cell(row=c, column=5).value = 1
                ws.cell(row=c, column=6).value = con_temp
                c += 1
                smart_count += 1
                # On the fourth loop set count to zero because the SMART set is complete (4 degenerate primer sets)
                if smart_count == 4:
                    smart_count = 0
            if values[6] == 'TRICK':
                ws.cell(row=c, column=7).value = values[forward_or_reverse]
                ws.cell(row=c, column=1).value = plate_count
                ws.cell(row=c, column=2).value = well_nums[int(c - 2)]
                con_temp = (mole_ratio_trick[trick_count] * f * (volume) / trick_volumes[trick_count])
                ws.cell(row=c, column=3).value = 100 - (con_temp * 100 / 5)
                ws.cell(row=c, column=4).value = 'water'
                ws.cell(row=c, column=5).value = 1
                ws.cell(row=c, column=6).value = con_temp
                c += 1
                trick_count += 1
                # On the fourth loop set count to zero because the TRICK set is complete (3 degenerate primer sets)
                if trick_count == 3:
                    trick_count = 0
            if well_nums[int(c - 2)-1] == 'H12':
                wb.remove(wb['Sheet'])
                wb.save(paths['output'] / f'Biomek_transfers_water_plate_{plate_count}.xlsx')
                save_data_to_csv(paths['output'] / f'Biomek_transfers_water_plate_{plate_count}.xlsx',
                                 paths['output'] / f'Biomek_transfers_water_plate_{plate_count}.csv')
                plate_count += 2
                wb = openpyxl.Workbook()
                ws = wb.create_sheet()
                ws.title = f"Plate {plate_count}"
                add_names_to_excel_sheet(ws)
                c = 2
            # if (plate_count == 7 or plate_count == 8) and c == 98:
            #     print("WARNING: Too many primers; output files likely have errors")
        wb.remove(wb['Sheet'])
        wb.save(paths['output'] / f'Biomek_transfers_water_plate_{plate_count}.xlsx')
        save_data_to_csv(paths['output'] / f'Biomek_transfers_water_plate_{plate_count}.xlsx',
                         paths['output'] / f'Biomek_transfers_water_plate_{plate_count}.csv')

def save_data_to_csv(file_name_in, file_name_out):
    xlsx = openpyxl.load_workbook(file_name_in)
    sheet = xlsx.active
    data = sheet.rows
    csv = open(file_name_out, "w+")
    for row in data:
        l = list(row)
        for i in range(len(l)):
            if i == len(l) - 1:
                csv.write(str(l[i].value))
            else:
                csv.write(str(l[i].value) + ',')
        csv.write('\n')
    csv.close()
    os.remove(file_name_in)

def add_names_to_excel_sheet(sheet):
    sheet.cell(row=1, column=1).value = 'Family ID'
    sheet.cell(row=1, column=2).value = 'Well ID'
    sheet.cell(row=1, column=3).value = 'Volume'
    sheet.cell(row=1, column=4).value = 'Reservoir Number'
    sheet.cell(row=1, column=5).value = 'Reservoir Well'
    sheet.cell(row=1, column=6).value = 'Concentration'
    sheet.cell(row=1, column=7).value = 'Primer ID'

def output_biomek_excel_file_all_primers(paths, assembly, volume, f):
    ordered_assembly = collections.OrderedDict(sorted(assembly.items()))
    mole_ratio_smart = [(12 / 20), (6 / 20), (1 / 20), (1 / 20)]
    mole_ratio_trick = [(12 / 22), (9 / 22), (1 / 22)]
    smart_volumes = [.5*(volume/20), .3*(volume/20), .1*(volume/20), .1*(volume/20)]
    trick_volumes = [.545454*(volume/20), .454545*(volume/20), .11363635*(volume/20)]
    well_nums_96 = ['ABCDEFGH'[idx // 12] + str(idx % 12 + 1) for idx in range(96)]
    well_nums = well_nums_96
    for x in range(int(math.ceil(len(ordered_assembly)/96))):
        well_nums = well_nums + well_nums_96
    wb = openpyxl.Workbook()
    ws = wb['Sheet']
    add_names_to_excel_sheet(ws)
    c = 2
    for forward_or_reverse in [1, 2]:
        smart_count = 0
        trick_count = 0
        plate_count = forward_or_reverse
        well_num_count = 0
        for key, values in ordered_assembly.items():
            if values[6] in ['Single', 'Degenerate']:
                # Add to a list based on counter
                ws.cell(row=c, column=7).value = values[forward_or_reverse]
                ws.cell(row=c, column=1).value = plate_count
                ws.cell(row=c, column=2).value = well_nums[well_num_count]
                ws.cell(row=c, column=3).value = (f / 0.05)*200/5
                ws.cell(row=c, column=4).value = 'Intermediate_Plate'
                ws.cell(row=c, column=5).value = well_nums[well_num_count]
                ws.cell(row=c, column=6).value = (f / 0.05)
                c += 1
                well_num_count += 1
            if values[6] == 'SMART':
                ws.cell(row=c, column=7).value = values[forward_or_reverse]
                ws.cell(row=c, column=1).value = plate_count
                ws.cell(row=c, column=2).value = well_nums[well_num_count]
                con_temp = (mole_ratio_smart[smart_count] * f * (volume) / smart_volumes[smart_count])
                ws.cell(row=c, column=3).value = con_temp * 200 / 5
                ws.cell(row=c, column=4).value = 'Intermediate_Plate'
                ws.cell(row=c, column=5).value = well_nums[well_num_count]
                ws.cell(row=c, column=6).value = con_temp
                c += 1
                well_num_count += 1
                smart_count += 1
                # On the fourth loop set count to zero because the SMART set is complete (4 degenerate primer sets)
                if smart_count == 4:
                    smart_count = 0
            if values[6] == 'TRICK':
                ws.cell(row=c, column=7).value = values[forward_or_reverse]
                ws.cell(row=c, column=1).value = plate_count
                ws.cell(row=c, column=2).value = well_nums[well_num_count]
                con_temp = (mole_ratio_trick[trick_count] * f * (volume) / trick_volumes[trick_count])
                ws.cell(row=c, column=3).value = con_temp * 200 / 5
                ws.cell(row=c, column=4).value = 'Intermediate_Plate'
                ws.cell(row=c, column=5).value = well_nums[well_num_count]
                ws.cell(row=c, column=6).value = con_temp
                c += 1
                well_num_count += 1
                trick_count += 1
                # On the fourth loop set count to zero because the TRICK set is complete (3 degenerate primer sets)
                if trick_count == 3:
                    trick_count = 0
            if well_nums[well_num_count-1] == 'H12':
                plate_count += 2

    wb.save(paths['output'] / f'SAMI_transfers_primers.xlsx')
    save_data_to_csv(paths['output'] / f'SAMI_transfers_primers.xlsx',
                     paths['output'] / f'SAMI_transfers_primers.csv')

    wb = openpyxl.Workbook()
    ws = wb['Sheet']
    ws.cell(row=1, column=1).value = 'Family ID'
    ws.cell(row=1, column=2).value = 'Well ID'
    ws.cell(row=1, column=3).value = 'Volume'
    ws.cell(row=1, column=4).value = 'Reservoir Number'
    ws.cell(row=1, column=5).value = 'Reservoir Well'

    c = 2
    for forward_or_reverse in [1, 2]:
        smart_count = 0
        trick_count = 0
        plate_count = forward_or_reverse
        well_num_count = 0
        for key, values in ordered_assembly.items():
            if values[6] in ['Single', 'Degenerate']:
                # Add to a list based on counter
                ws.cell(row=c, column=1).value = plate_count
                ws.cell(row=c, column=2).value = well_nums[well_num_count]
                ws.cell(row=c, column=3).value = 200 - ((f / 0.05) * 200 / 5)
                ws.cell(row=c, column=4).value = 'water'
                ws.cell(row=c, column=5).value = 1
                c += 1
                well_num_count += 1
            if values[6] == 'SMART':
                ws.cell(row=c, column=1).value = plate_count
                ws.cell(row=c, column=2).value = well_nums[well_num_count]
                con_temp = (mole_ratio_smart[smart_count] * f * (volume) / smart_volumes[smart_count])
                ws.cell(row=c, column=3).value = 200-(con_temp * 200 / 5)
                ws.cell(row=c, column=4).value = 'water'
                ws.cell(row=c, column=5).value = 1
                c += 1
                well_num_count += 1
                smart_count += 1
                # On the fourth loop set count to zero because the SMART set is complete (4 degenerate primer sets)
                if smart_count == 4:
                    smart_count = 0
            if values[6] == 'TRICK':
                ws.cell(row=c, column=1).value = plate_count
                ws.cell(row=c, column=2).value = well_nums[well_num_count]
                con_temp = (mole_ratio_trick[trick_count] * f * (volume) / trick_volumes[trick_count])
                ws.cell(row=c, column=3).value = 200 - (con_temp * 200 / 5)
                ws.cell(row=c, column=4).value = 'water'
                ws.cell(row=c, column=5).value = 1
                c += 1
                well_num_count += 1
                trick_count += 1
                # On the fourth loop set count to zero because the TRICK set is complete (3 degenerate primer sets)
                if trick_count == 3:
                    trick_count = 0
            if well_nums[well_num_count-1] == 'H12':
                plate_count += 2

    wb.save(paths['output'] / f'SAMI_transfers_water.xlsx')
    save_data_to_csv(paths['output'] / f'SAMI_transfers_water.xlsx',
                     paths['output'] / f'SAMI_transfers_water.csv')
