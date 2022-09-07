import openpyxl
import re


def output_idt(file_name, primers, missing_primers):
    # group primers by plate
    n = 96
    # Old plates code to make plates with mixed forward and reverse primers
    # plates = [primers[i * n:(i + 1) * n] for i in range((len(primers) + n - 1) // n)]

    # New plates code to make plates that have segregated forward and reverse primers
    primers_temp = [primers[::2], primers[1::2]]
    f_primers = [primers_temp[0][i:i + n] for i in range(0, len(primers_temp[0]), n)]
    r_primers = [primers_temp[1][i:i + n] for i in range(0, len(primers_temp[1]), n)]
    plates = []
    for f, r in zip(f_primers, r_primers):
        plates.append(f)
        plates.append(r)
    # print(plates)

    # create workbook
    wb = openpyxl.Workbook()

    # create a sheet per plate and add the info
    for plate_num, plate in enumerate(plates, start=1):
        # print(plate_num)
        sheet = wb.create_sheet()
        sheet.title = f"Plate {plate_num}"

        # Add header to sheet
        row = 1
        a1 = sheet.cell(row=row, column=1)
        a1.value = 'Well'
        b1 = sheet.cell(row=row, column=2)
        b1.value = 'Name'
        c1 = sheet.cell(row=row, column=3)
        c1.value = 'Sequence'
        row = 2
        # Add the well, primer and sequence to the sheet
        for idx, primer in enumerate(plate, start=1):
            well = sheet.cell(row=row, column=1)
            well_num = 'ABCDEFGH'[(idx - 1) // 12] + '%02d' % ((idx - 1) % 12 + 1,)
            well.value = str(well_num)

            primer_name = sheet.cell(row=row, column=2)
            primer_name.value = primer[0]

            primer_sequence = sheet.cell(row=row, column=3)
            primer_sequence.value = str(primer[1])
            row = row + 1

    print('Number of plates: ' + str(plate_num))
    # remove default sheet
    sheet = wb['Sheet']
    wb.remove(sheet)

    # Add sheet with list of missing mutations
    sheet = wb.create_sheet()
    sheet.title = "Missing Primers"
    a1 = sheet.cell(row=1, column=1)
    a1.value = 'Target Mutation'
    for idx, mutation in enumerate(missing_primers, start=2):
        sheet.cell(row=idx, column=1).value = mutation

    # write excel file
    wb.save(f"{str(file_name).split('.')[0]}.xlsx")


def read_input_file(filename):
    primers = list()
    with open(filename) as myFile:
        for line in myFile:
            primer_name_and_sequence = re.split(r'\t', line)
            primers.append(primer_name_and_sequence)  # add primer to list of primers to write to file
    return primers
