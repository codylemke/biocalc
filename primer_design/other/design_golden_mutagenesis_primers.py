import math

import numpy as np
import openpyxl
import pandas as pd
import rpy2
print(rpy2.__version__)
# from rpy2.rinterface import R_VERSION_BUILD
# print(R_VERSION_BUILD)
import rpy2.robjects as robjects

# https://stackoverflow.com/questions/49485607/installing-custom-r-package-from-bitbucket-in-python-through-rpy2
from rpy2.robjects.packages import importr
import rpy2.robjects.packages as rpackages
from rpy2.robjects.vectors import StrVector, BoolVector
utils = rpackages.importr('utils')
# Devtools and GoldenMutagenesis installation code
# devtools = utils.install_packages('devtools')
# d = {'package.dependencies': 'package_dot_dependencies',
#      'package_dependencies': 'package_uscore_dependencies'}
# custom_analytics = importr('devtools', robject_translations = d)
# golden = custom_analytics.install_github("https://github.com/ipb-halle/GoldenMutagenesis.git")

from primer_design import design_quickchange_primers
from pathlib import Path
from os import chdir, listdir, system
import os
import glob
from argparse import Namespace
from openpyxl import load_workbook
from Bio import SeqIO
import sys
from primer_design import idt_primer_output

def check_if_directory_exists_and_make_directory(directory):
    if directory.exists():  # check if experiment id exists, otherwise make directory
        pass
    else:
        directory.mkdir()


def prompt_with_directory_files(directory, text):
    files = list()
    print("----------------")
    for idx, file in enumerate(sorted(directory.rglob('*')), start=1):
        print(idx, "-", file.name)
        files.append(file)
    selected_file_number = int(input(f"{text}"))
    return files[selected_file_number-1]


def setup_experiment_directories():
    # Change directory to Experiments folder
    os.chdir('../')
    path = Path.cwd() / "Experiments"
    chdir(path)
    # print(Path.cwd())
    dir_list = []
    for idx, dir_name in enumerate(sorted([name for name in os.listdir(".") if os.path.isdir(name)]), start=1):
        print(idx, "-", dir_name)
        dir_list.append(dir_name)
    # Get experiment ID from user, create new directory structure for experiment, move to new directory
    selected_id = int(input("Please enter the number associated with the directory with your files:"))
    experiment_id = dir_list[selected_id-1]
    path = Path.cwd() / f'{experiment_id}'

    if path.exists():  # check if experiment id exists, otherwise make directory
        pass
    else:
        path.mkdir()

    chdir(path)  # move to new directory

    # create input and output directories if needed
    input_path = Path.cwd() / 'input'
    output_path = Path.cwd() / 'output'
    check_if_directory_exists_and_make_directory(input_path)
    check_if_directory_exists_and_make_directory(output_path)

    experiment_paths = {"input": input_path, 'output': output_path, "experiment_id": experiment_id}
    return experiment_paths


def make_primers(params):
    # Load R Functions
    GM = rpackages.importr("GoldenMutagenesis")
    r_list = robjects.r['list']
    robjects.r('''
              access_primer_oldseq <- function(primers){
              primers@oldsequence}
              ''')
    robjects.r('''
              access_primer_newseq <- function(primers){
              primers@newsequence}
              ''')
    robjects.r('''
              primers_obj_len <- function(primers){
              length(primers@fragments)}
              ''')
    robjects.r('''
              access_primer_fragments <- function(primers, idx1, name){
              slot(primers@fragments[[idx1]], name)}
              ''')
    robjects.r('''
              access_primer_primers <- function(primers, idx1, idx2, name){
              slot(primers@primers[[idx1]][[idx2]], name)}
              ''')
    r_primers_obj_len = robjects.r['primers_obj_len']
    r_access_primer_fragments = robjects.r['access_primer_fragments']
    r_access_primer_primers = robjects.r['access_primer_primers']
    r_access_primer_oldseq = robjects.r['access_primer_oldseq']
    r_access_primer_newseq = robjects.r['access_primer_newseq']

    seqs = list(SeqIO.parse(params.i, "fasta"))
    if len(seqs) != 1:
        raise ValueError("Error: Input fasta must contain only one sequence")
    seq_name = seqs[0].name
    seq = seqs[0].seq
    seq_str = str(seq)
    mod_val = len(seq_str) % 3
    print(len(seq_str))
    if mod_val != 0:
        input_sequence = seq_str[:-mod_val]
        print("Length of Sequence not divisable by 3. Removed last:", mod_val, "base(s)")
        print("Input sequence length:", len(input_sequence))
    else:
        input_sequence = seq_str

    df = pd.read_excel(params.target_mutations, sheet_name='Mutations')  # r"..\Experiments\Golden_testing\mutations.xlsx"

    locations = df["Location"].tolist()
    mutation_codes = df["Mutation"].tolist()
    mut_ids = df["ID"].tolist()

    primer_ids = []
    rxn_id = []
    primer_seqs = []
    primer_codes = []
    temps = []
    fragment_lengths = []
    fragment_ids = []
    fragment_rxn_id = []
    fragment_number = []
    primer_number = []
    primer_direction = []

    if params.restriction_enzyme == "bsai":
        restrict_enzyme = params.recognition_site_bsai
        suffix_input = params.suffix  # "A"  # Make param?
    elif params.restriction_enzyme == "bbsi":
        restrict_enzyme = params.recognition_site_bbsi
        suffix_input = params.suffix + params.suffix  #"AA"  # Make param?
    else:
        print('ERROR: Restriction enzyme ID is incorrect.')

    for location, mut_code, mut_id in zip(locations, mutation_codes, mut_ids):
        mutations = r_list(location)
        primers = GM.mutate_msd(input_sequence, prefix="TT",
                                restriction_enzyme=restrict_enzyme,
                                codon=params.codon,
                                suffix=suffix_input, vector=robjects.StrVector(["AATG", "CACT"]),  # 5 -> 3 on each strand
                                replacements=mutations[0], replacement_range=5,
                                binding_min_length=4, binding_max_length=9,
                                target_temp=60, fragment_min_size=31)  #31

        num_primers = r_primers_obj_len(primers)[0]
        bool_type = type(r_access_primer_fragments(primers, 1, 'start_mutation'))

        for i in range(1, num_primers+1):
            # for name in ['start', 'stop', 'start_mutation', 'stop_mutation']:
            fragment_length = (r_access_primer_fragments(primers, i, 'stop')[0] -
                                r_access_primer_fragments(primers, i, 'start')[0] - 1) * 3
            fragment_lengths.append(fragment_length)
            fragment_ids.append(str(mut_id) + '_' + str(i))
            fragment_rxn_id.append(str(mut_id))
            fragment_number.append(i)

            for j in [1, 2]:  # Forward then reverse
                primer_seq_temp = ''
                for name in ['prefix', 'restriction_enzyme', 'suffix', 'vector', 'overhang', 'extra', 'binding_sequence']:
                    primer_seq_temp = primer_seq_temp + r_access_primer_primers(primers, i, j, name)[0]
                if j == 1:
                    # print('Forward:', primer_seq_temp)
                    primer_ids.append(str(mut_id)+'_'+str(i)+'_F')
                    rxn_id.append(str(mut_id))
                    primer_number.append(i)
                    primer_direction.append('F')
                    primer_seqs.append(primer_seq_temp)
                    primer_codes.append(mut_code)
                    temps.append(r_access_primer_primers(primers, i, j, 'temperature')[0])
                else:
                    # print('Reverse:', primer_seq_temp)
                    primer_ids.append(str(mut_id) + '_' + str(i) + '_R')
                    rxn_id.append(str(mut_id))
                    primer_number.append(i)
                    primer_direction.append('R')
                    primer_seqs.append(primer_seq_temp)
                    primer_codes.append(mut_code)
                    temps.append(r_access_primer_primers(primers, i, j, 'temperature')[0])
                # for name2 in ['temperature', 'difference']:
                #     print(r_access_primer_primers(primers, i, j, name2)[0])

    output_df = pd.DataFrame.from_dict(
        {"Primer_ID": primer_ids, "Rxn_ID": rxn_id, "Fragment_num": primer_number, "Direction": primer_direction,
         "Sequence": primer_seqs, "Mutation": primer_codes, "Temp": temps})
    output_df.to_excel(params.o)

    fragments_df = pd.DataFrame.from_dict({"ID": fragment_ids, "Rxn_ID": fragment_rxn_id,
                                           "Fragment_num": fragment_number, "Length_nt": fragment_lengths})
    fragments_df.to_excel(params.paths['output'] / 'fragment_info.xlsx')

    return output_df, fragments_df


def make_mutagenesis_primers(directory):
    print('here')

    gene_sequence_directory = prompt_with_directory_files(directory["input"], "Please enter the number associated with the gene sequence file:")
    mutagenesis_list = prompt_with_directory_files(directory["input"], "Please enter the number associated with the mutagenesis file:" )

    primers_output = directory['output'] / 'OG_primer_library.xlsx'
    assembly_file_out = directory['output'] / 'new_assembly_golden.tsv'

    wb = load_workbook(mutagenesis_list, data_only=True)
    sh = wb["Inputs"]

    args = Namespace(
                     i=gene_sequence_directory,
                     o=primers_output,
                     paths=directory,
                     output_assembly_file=assembly_file_out,
                     target_mutations=mutagenesis_list,
                     temp_conc=sh["b2"].value,
                     primer_conc=sh["b3"].value,
                     rxn_vol=sh["b4"].value,
                     stock_volume=200,
                     prefix=sh["b6"].value,
                     restriction_enzyme=sh["b7"].value,
                     codon=sh["b8"].value,
                     suffix=sh["b9"].value,
                     replace_range=sh["b10"].value,
                     min_length=sh["b11"].value,
                     max_length=sh["b12"].value,
                     target_temp=sh["b13"].value,
                     trick=sh["b14"].value,
                     quikchange=True,
                     recognition_site_bbsi="GAAGAC",
                     recognition_site_bsai="GGTCTC",
                     target_species_taxid=0,  # sh["b15"].value, #316407,
                     )

    output_df, fragments_df = make_primers(args)
    return output_df, fragments_df, args


def convert_primers_to_trick(params, primers_df):
    keys = []
    values = []
    rxn_id = []
    primer_codes = []
    temps = []
    primer_fragment_number = []
    primer_direction = []
    trick_id = []
    primer_length = []
    for index, row in primers_df.iterrows():
        seq = row['Sequence']
        if params.codon in seq:
            for code, i in zip(['NDT', 'VHG', 'TGG'], ['a', 'b', 'c']):
                trick_seq = seq.replace(params.codon, code)
                values.append(row['Primer_ID']+i)
                keys.append(trick_seq)
                trick_id.append(i)
                rxn_id.append(row['Rxn_ID'])
                primer_codes.append(row['Mutation'][0:-5]+'('+code+')')
                temps.append(row['Temp'])
                primer_fragment_number.append(row['Fragment_num'])
                primer_direction.append(row['Direction'])
                primer_length.append(len(trick_seq))
        elif params.codon[::-1] in seq:
            for code, i in zip(['NHA', 'BDC', 'ACC'], ['a', 'b', 'c']):
                trick_seq = seq.replace(params.codon[::-1], code[::-1])
                values.append(row['Primer_ID']+i)
                keys.append(trick_seq)
                trick_id.append(i)
                rxn_id.append(row['Rxn_ID'])
                primer_codes.append(row['Mutation'][0:-5]+'('+code+')')
                temps.append(row['Temp'])
                primer_fragment_number.append(row['Fragment_num'])
                primer_direction.append(row['Direction'])
                primer_length.append(len(trick_seq))
        else:
            values.append(row['Primer_ID']+'o')
            keys.append(seq)
            trick_id.append('o')
            rxn_id.append(row['Rxn_ID'])
            primer_codes.append(row['Mutation'][0:-5])
            temps.append(row['Temp'])
            primer_fragment_number.append(row['Fragment_num'])
            primer_direction.append(row['Direction'])
            primer_length.append(len(seq))

    all_trick_primers_df = pd.DataFrame({'IDT_ID': values, 'Primer_ID': values, "Rxn_ID": rxn_id,
                                         "Fragment_num": primer_fragment_number,
                                         'Trick_ID': trick_id, 'Direction': primer_direction, 'Sequence': keys,
                                         'Mutation': primer_codes, 'Temp': temps, 'Length': primer_length})

    trick_primers_dict = {}
    for key, value in zip(keys, values):
        trick_primers_dict.setdefault(key, []).append(value)
    print("Total Primers:", len(keys), "Unique Primers:", len(trick_primers_dict))

    uses = []
    for index, row in all_trick_primers_df.iterrows():
        num_rxns = len(trick_primers_dict[row['Sequence']])
        uses.append(num_rxns)
    all_trick_primers_df['Rxn_Uses'] = uses

    all_trick_primers_df.to_excel(params.paths['output'] / 'all_trick_primers.xlsx')

    return all_trick_primers_df  # , trick_primers_df


def generate_fragment_info(params, fragments_df, all_trick_primers_df):
    well_ids = list(range(1, 385)) + list(range(1, 385)) + list(range(1, 385)) + list(range(1, 385))
    all_parts = []
    plate_id = []
    well_id = []
    for index, row in fragments_df.iterrows():
        fragment_id = row['Rxn_ID']
        fragment_number = row['Fragment_num']
        parts = []
        for index_temp, primer_row in all_trick_primers_df.loc[all_trick_primers_df['Rxn_ID'] ==
                                                               str(fragment_id)].iterrows():
            if primer_row['Fragment_num'] == fragment_number:
                parts.append(primer_row['Primer_ID'])
        all_parts.append(parts)

        plate_id.append(1 + math.floor((index+1)/385))
        well_id.append(well_ids[index])

    fragments_df['Parts_IDs'] = all_parts
    fragments_df['Plate_IDs'] = plate_id
    fragments_df['Well_IDs'] = well_id

    fragments_df.to_excel(params.paths['output'] / 'fragments_df.xlsx')
    return fragments_df


def part_locations_and_concentrations(params, fragments_df, all_trick_primers_df):

    # Part plates (including Plate <0>)
    # Plate <0> part wells
    well_nums = ['ABCDEFGH'[idx // 12] + str(idx % 12 + 1) for idx in range(96)]
    trick_primers_dict = {}
    for key, value in zip(all_trick_primers_df['Sequence'], all_trick_primers_df['Primer_ID']):
        trick_primers_dict.setdefault(key, []).append(value)

    IDT_IDs = []
    keys_list_sp = []
    num_rxns_list_sp = []
    primer_ids_sp = []
    all_well_ids = []
    plate_ids = []
    plate_nums = []
    count = 1
    num_rxns_total = len(fragments_df)
    well_count = 0 + 3 + math.ceil(num_rxns_total/180)
    plate_num = 0
    for key, value in trick_primers_dict.items():
        if len(key) > 60 or len(value) > 1:
            IDT_IDs.append('IDT-'+str(count))
            keys_list_sp.append(key)
            num_rxns_list_sp.append(len(value))
            primer_ids_sp.append(value)
            count += 1
            well_ids = []
            iter_range = range(math.ceil(len(value)/180))
            if well_count + len(iter_range) > 96:
                plate_num += 1
                well_count = 0
            for x in iter_range:
                well_ids.append(well_nums[well_count])
                well_count += 1
            plate_ids.append('Plate <' + str(plate_num) + '>')
            all_well_ids.append(well_ids)

    manual_trick_primers_df = pd.DataFrame({'IDT_ID': IDT_IDs, 'Primer': keys_list_sp,
                                            'Rxns required': num_rxns_list_sp, 'IDs': primer_ids_sp,
                                            'Plate_ID': plate_ids, 'Well_IDs': all_well_ids})


    for index, row in all_trick_primers_df.iterrows():
        for index2, man_row in manual_trick_primers_df.iterrows():
            if row['Primer_ID'] in man_row['IDs']:
                all_trick_primers_df.at[index, 'IDT_ID'] = man_row['IDT_ID']

    # Source concentration and Final concentration
    concentration_dict = {'o': params.primer_conc, 'a': params.primer_conc*(12/22),
                          'b': params.primer_conc*(9/22), 'c': params.primer_conc*(1/22)}
    trick_volumes = {'o': 1, 'a': .545454, 'b': .454545, 'c': .11363635}
    final_conc = []
    input_conc = []
    plate_ID = []
    plate_nums = []
    well_ID = []
    biomek_xfer_vol = []
    biomek_input_conc = []
    manual_conc_dict = {}
    count = 0
    plate_num += 1  # Start on next plate number

    # Calculate the concentrations required for appropriate volume transfer on the Nanohive
    for index, row in all_trick_primers_df.iterrows():
        final_conc.append(concentration_dict[row['Trick_ID']])
        concentration_from_biomek = concentration_dict[row['Trick_ID']]*(params.rxn_vol/trick_volumes[row['Trick_ID']])
        input_conc.append(concentration_from_biomek)

        if row['Length'] > 60 or row['Rxn_Uses'] > 1:
            plate_ID.append('Plate <Manual>')
            plate_nums.append(np.nan)
            well_ID.append('tbd')
            biomek_xfer_vol.append(np.nan)
            biomek_input_conc.append(np.nan)
            manual_conc_dict[row['IDT_ID']] = concentration_from_biomek
        else:
            well_ID.append(well_nums[count])
            plate_ID.append('Plate <' + str(plate_num) + '>')
            plate_nums.append(plate_num)
            biomek_xfer_vol.append(concentration_from_biomek*200/5)  # C2*V2/C1 [uM and uL]
            biomek_input_conc.append(5)  # uM
            count += 1
            if count == 96:
                plate_num += 1
                count = 0

    all_trick_primers_df['Plate_IDs'] = plate_ID
    all_trick_primers_df['Plate_num'] = plate_nums
    all_trick_primers_df['Well_IDs'] = well_ID
    all_trick_primers_df['Final[uM]'] = final_conc
    all_trick_primers_df['Input[uM]'] = input_conc
    all_trick_primers_df['Bio_In[uM]'] = biomek_input_conc
    all_trick_primers_df['Bio_vol[uL]'] = biomek_xfer_vol

    all_trick_primers_df.to_excel(params.paths['output'] / 'all_trick_primers.xlsx')

    concentrations_list = []
    for index, row in manual_trick_primers_df.iterrows():
        key = row['IDT_ID']
        concentrations_list.append(manual_conc_dict[key])

    manual_trick_primers_df['Input[uM]'] = concentrations_list
    manual_trick_primers_df.to_excel(params.paths['output'] / 'manual_trick_primers.xlsx')

    return all_trick_primers_df, manual_trick_primers_df


def IDT_order_files(all_trick_primers_df):
    well = []
    name = []
    seq = []
    file_name = 'IDT_order.xlsx'
    with pd.ExcelWriter(params.paths['output'] / file_name) as writer:
        for index, row in all_trick_primers_df.iterrows():
            if row['Length'] <= 60 and row['Rxn_Uses'] == 1:
                well.append(row['Well_IDs'])
                name.append(row['IDT_ID'])
                seq.append(row['Sequence'])
                plate_id = str(int(row['Plate_num']))
                if row['Well_IDs'] == 'H12':
                    temp_df = pd.DataFrame({'Well Position': well, 'Name': name, 'Sequence': seq})
                    temp_df.to_excel(writer, sheet_name='Plate <'+plate_id+'>', index=False)
                    well = []
                    name = []
                    seq = []
        temp_df = pd.DataFrame({'Well Position': well, 'Name': name, 'Sequence': seq})
        temp_df.to_excel(writer, sheet_name='Plate <'+plate_id+'>', index=False)


def Nanohive_assembly_file(params, fragments_df, manual_trick_primers_df, all_trick_primers_df):
    template_file_path = r'..\Nanohive\Assembly Template GM.xlsx'
    wb = load_workbook(filename=template_file_path)

    for plate_num in range(2, int(max(all_trick_primers_df['Plate_num']))+1):
        wb.copy_worksheet(wb['Plate <0>']).title = 'Plate <' + str(plate_num) + '>'
        wb.save(paths['output'] / 'Nanohive_Assembly.xlsx')

    excel_row = 3
    # plate_count = 2
    for index, row in all_trick_primers_df.iterrows():
        if row['Length'] <= 60 and row['Rxn_Uses'] == 1:
            ws = wb[row['Plate_IDs']]
            ws.cell(row=excel_row, column=2).value = row['Primer_ID']
            ws.cell(row=excel_row, column=5).value = round(row['Input[uM]'], 3)
            ws.cell(row=excel_row, column=6).value = 'Concentration (µM)'
            ws.cell(row=excel_row, column=7).value = params.stock_volume
            excel_row += 1
            if excel_row == 99:
                excel_row = 3
                # wb.copy_worksheet(wb['Plate <0>']).title = 'Plate <'+str(plate_count)+'>'
                # plate_count += 1
                # wb.save(paths['output'] / 'Nanohive_Assembly.xlsx')

    # Add vectors and controls to plate zero
    ws = wb['Plate <0>']
    ws.cell(row=3, column=2).value = 'Control_F'
    ws.cell(row=3, column=5).value = params.primer_conc
    ws.cell(row=3, column=6).value = 'Concentration (µM)'
    ws.cell(row=3, column=7).value = params.stock_volume

    ws.cell(row=4, column=2).value = 'Control_R'
    ws.cell(row=4, column=5).value = params.primer_conc
    ws.cell(row=4, column=6).value = 'Concentration (µM)'
    ws.cell(row=4, column=7).value = params.stock_volume

    ws.cell(row=5, column=2).value = 'Control_construct'
    ws.cell(row=5, column=5).value = params.temp_conc
    ws.cell(row=5, column=6).value = 'Concentration (ng/µL)'
    ws.cell(row=5, column=7).value = params.stock_volume - 20

    construct_well_count = int(math.ceil(len(fragments_df)/(params.stock_volume - 20)))

    for c in range(6, 6+construct_well_count):
        ws.cell(row=c, column=2).value = 'Exp_plasmid'  # Get construct name
        ws.cell(row=c, column=5).value = params.temp_conc * 20
        ws.cell(row=c, column=6).value = 'Concentration (ng/µL)'
        ws.cell(row=c, column=7).value = params.stock_volume - 20

    current_plate_id = 'Plate <0>'
    for index, row in manual_trick_primers_df.iterrows():
        for well_id in row['Well_IDs']:
            if row['Plate_ID'] != current_plate_id:
                c = 3
                ws = wb[row['Plate_ID']]
            else:
                c += 1
            ws.cell(row=c, column=2).value = row['IDT_ID']
            ws.cell(row=c, column=5).value = row['Input[uM]']
            ws.cell(row=c, column=6).value = 'Concentration (µM)'
            ws.cell(row=c, column=7).value = params.stock_volume
            current_plate_id = row['Plate_ID']

    wb.save(paths['output'] / 'Nanohive_Assembly.xlsx')

    ws = wb['Assembly List']
    c = 2
    for index, row in fragments_df.iterrows():
        ws.cell(row=c, column=4).value = row['ID']
        ws.cell(row=c, column=5).value = params.rxn_vol
        ws.cell(row=c, column=6).value = 'Exp_plasmid'
        ws.cell(row=c, column=7).value = params.temp_conc
        ws.cell(row=c, column=8).value = 'Concentration (ng/µL)'
        c += 1
        for part in row['Parts_IDs']:
            ws.cell(row=c, column=4).value = row['ID']
            ws.cell(row=c, column=5).value = params.rxn_vol
            ws.cell(row=c, column=6).value = part
            concentration_row = all_trick_primers_df.loc[all_trick_primers_df['Primer_ID'] == part]
            concentration = concentration_row['Final[uM]'].iloc[0]
            # print(concentration)
            ws.cell(row=c, column=7).value = concentration
            ws.cell(row=c, column=8).value = 'Concentration (µM)'
            c += 1

    wb.save(paths['output'] / 'Nanohive_Assembly.xlsx')


def biomek_files(all_trick_primers_df):
    well = []
    name = []
    family_id = []
    volume = []
    res_num = []
    concen = []
    for index, row in all_trick_primers_df.iterrows():
        if row['Length'] <= 60 and row['Rxn_Uses'] == 1:
            well.append(row['Well_IDs'])
            name.append(row['IDT_ID'])
            plate_id = int(row['Plate_num'])
            family_id.append(plate_id)
            volume.append(round(row['Bio_vol[uL]'],3))
            res_num.append('Intermediate_Plate')
            concen.append(row['Bio_In[uM]'])
            if row['Well_IDs'] == 'H12':
                temp_df = pd.DataFrame({'Family ID': family_id, 'Well ID': well, 'Volume': volume,
                                        'Reservoir Number': res_num, 'Reservoir Well': well,
                                        'Concentration': concen, 'Primer ID': name})
                temp_df.to_excel(paths['output'] / ('Biomek_Plate_'+str(plate_id)+'.xlsx'), sheet_name='Transfers', index=False)
                save_data_to_csv(paths['output'] / ('Biomek_Plate_' + str(plate_id) + '.xlsx'),
                                 paths['output'] / ('Biomek_Plate_' + str(plate_id) + '.csv'))
                well = []
                name = []
                family_id = []
                volume = []
                res_num = []
                concen = []
    temp_df = pd.DataFrame({'Family ID': family_id, 'Well ID': well, 'Volume': volume,
                            'Reservoir Number': res_num, 'Reservoir Well': well,
                            'Concentration': concen, 'Primer ID': name})
    temp_df.to_excel(paths['output'] / ('Biomek_Plate_'+str(plate_id)+'.xlsx'), sheet_name='Transfers', index=False)
    save_data_to_csv(paths['output'] / ('Biomek_Plate_'+str(plate_id)+'.xlsx'),
                     paths['output'] / ('Biomek_Plate_'+str(plate_id)+'.csv'))

    well = []
    name = []
    family_id = []
    volume = []
    res_num = []
    concen = []
    for index, row in all_trick_primers_df.iterrows():
        if row['Length'] <= 60 and row['Rxn_Uses'] == 1:
            well.append(row['Well_IDs'])
            name.append(row['IDT_ID'])
            plate_id = int(row['Plate_num'])
            family_id.append(plate_id)
            volume.append(200 - round(row['Bio_vol[uL]'], 3))
            res_num.append('Intermediate_Plate')
            concen.append(row['Bio_In[uM]'])
            if row['Well_IDs'] == 'H12':
                temp_df = pd.DataFrame({'Family ID': family_id, 'Well ID': well, 'Volume': volume,
                                        'Reservoir Number': res_num, 'Reservoir Well': well,
                                        'Concentration': concen, 'Primer ID': name})
                temp_df.to_excel(paths['output'] / ('Biomek_H2O_Plate_'+str(plate_id)+'.xlsx'), sheet_name='Transfers', index=False)
                save_data_to_csv(paths['output'] / ('Biomek_H2O_Plate_' + str(plate_id) + '.xlsx'),
                                 paths['output'] / ('Biomek_H2O_Plate_' + str(plate_id) + '.csv'))
                well = []
                name = []
                family_id = []
                volume = []
                res_num = []
                concen = []
    temp_df = pd.DataFrame({'Family ID': family_id, 'Well ID': well, 'Volume': volume,
                            'Reservoir Number': res_num, 'Reservoir Well': well,
                            'Concentration': concen, 'Primer ID': name})
    temp_df.to_excel(paths['output'] / ('Biomek_H2O_Plate_'+str(plate_id)+'.xlsx'), sheet_name='Transfers', index=False)
    save_data_to_csv(paths['output'] / ('Biomek_H2O_Plate_'+str(plate_id)+'.xlsx'),
                     paths['output'] / ('Biomek_H2O_Plate_'+str(plate_id)+'.csv'))

def save_data_to_csv(file_name_in, file_name_out):
    xlsx = openpyxl.load_workbook(file_name_in)
    sheet = xlsx.active
    data = sheet.rows
    csv = open(file_name_out, "w+")
    for row in data:
        l = list(row)
        for i in range(len(l)):
            if i == len(l) - 1:
                csv.write(str(l[i].value))
            else:
                csv.write(str(l[i].value) + ',')
        csv.write('\n')
    csv.close()
    os.remove(file_name_in)

if __name__ == "__main__":
    paths = setup_experiment_directories()
    primers_df, fragments_df, params = make_mutagenesis_primers(paths)
    if params.trick:
        if params.codon == params.codon[::-1]:
            sys.exit("ERROR: INPUT CODON MUST BE NON-PALINDROMIC FOR TRICK LIBRARY.")
        all_trick_primers_df = convert_primers_to_trick(params, primers_df)
        fragments_df = generate_fragment_info(params, fragments_df, all_trick_primers_df)
        all_trick_primers_df, manual_trick_primers_df = part_locations_and_concentrations(params, fragments_df,
                                                                                          all_trick_primers_df)
        IDT_order_files(all_trick_primers_df)
        Nanohive_assembly_file(params, fragments_df, manual_trick_primers_df, all_trick_primers_df)
        biomek_files(all_trick_primers_df)


        # Make biomek transfer files and rehydration of IDT plates --> 100 uM
        # Deal with difference in fragment length? (make two different plates based on length)
        # Output mutations that need QuikChange primers?
        # TODO: ADD TOTAL MASS INFORMATION ABOUT THE MANUAL PRIMERS FILE

## OLD CODE:

# cuf = "e_coli_316407.csv"
#
# mutations_bbsi = GM.domesticate(input_sequence, recognition_site_bbsi, cuf)
# mutations_bsai = GM.domesticate(input_sequence, recognition_site_bsai, cuf)
#
# first_param_temp = robjects.StrVector(["66", "V"])

# temp_var = r_list(first_param_temp)
# mutations = r_list(temp_var[0], mutations_bbsi[0])

# primers = GM.mutate_spm(input_sequence, prefix="TT", restriction_enzyme = recognition_site_bbsi,
#                         suffix = "AA", vector=robjects.StrVector(["CTCA", "CTCG"]), replacements = mutations,
#                         binding_min_length=4 ,binding_max_length = 9, target_temp=60,
#                         fragment_min_size = 60, cuf=cuf)

# print(r_access_primer_oldseq(primers)[0])
# print(type(r_access_primer_newseq(primers)[0]))

# help(bool_type)
# print([x for x in r_access_primer_fragments(primers, 3, 'start_mutation').items()])

# for i in range(1, num_primers+1):
#      for name in ['start', 'stop', 'start_mutation', 'stop_mutation']:
#           try:
#                print(r_access_primer_fragments(primers, i, name)[0])
#           except:
#                if type(r_access_primer_fragments(primers, i, name)) == bool_type:
#                     # print('pass')
#                     # print(type(r_access_primer_fragments(primers, i, name)))
#                     print([x for x in r_access_primer_fragments(primers, i, name).items()])
#                else:
#                     print('Warning: datatype issue')
#                     print(i, name)
#                     # temp_var = r_access_primer_fragments(primers, i, name)
#                     print(type(r_access_primer_fragments(primers, i, name)))