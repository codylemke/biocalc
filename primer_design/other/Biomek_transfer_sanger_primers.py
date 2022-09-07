import openpyxl
import os
import pandas as pd


def output_biomek_file(input_files, output_file):
    # import primer info
    primer_opts = pd.read_excel(input_files[0])
    # import well locations
    primer_info = pd.read_excel(input_files[1])
    # import primer-to-mutant table
    plate_layout = pd.read_excel(input_files[2], sheet_name='Data')

    # well_nums = ['ABCDEFGH'[idx // 12] + str(idx % 12 + 1) for idx in range(96)]

    c = 2
    wb = openpyxl.Workbook()
    ws = wb.create_sheet()
    # ws.title = f"Plate {sheet_name}"
    add_names_to_excel_sheet(ws)
    primer_usage_dict = {}
    for i, row in plate_layout.iterrows():
        primer_id = primer_opts.loc[primer_opts['ID'] == row['Strain Description']]['1st_Option'].values[0]
        if primer_id in list(primer_usage_dict.keys()):
            temp_count = primer_usage_dict[primer_id]
            primer_usage_dict[primer_id] = temp_count + 1
        else:
            primer_usage_dict[primer_id] = 1

    for i, row in plate_layout.iterrows():
        primer_id = primer_opts.loc[primer_opts['ID'] == row['Strain Description']]['1st_Option'].values[0]
        family_id = 'Plate1'
        well_id = row['Well']
        dest_plate_name = 'Primers'
        res_well = primer_info.loc[primer_info['Primer Number'] == primer_id]['Well ID'].values[0]

        ws.cell(row=c, column=1).value = family_id
        ws.cell(row=c, column=2).value = well_id
        ws.cell(row=c, column=3).value = 5
        ws.cell(row=c, column=4).value = dest_plate_name
        ws.cell(row=c, column=5).value = res_well
        ws.cell(row=c, column=6).value = 5
        ws.cell(row=c, column=7).value = primer_id
        ws.cell(row=c, column=8).value = 'At least: ' + str(primer_usage_dict[primer_id]*5)+' uL'
        ws.cell(row=c, column=9).value = row['Strain Description']
        ws.cell(row=c, column=10).value = row['Amino acid']

        c += 1
    wb.remove(wb['Sheet'])
    wb.save(output_file)
    save_data_to_csv(output_file, output_file.replace('.xlsx', '.csv'))

    print(primer_usage_dict)


def add_names_to_excel_sheet(sheet):
    sheet.cell(row=1, column=1).value = 'Family ID'
    sheet.cell(row=1, column=2).value = 'Well ID'
    sheet.cell(row=1, column=3).value = 'Volume uL'
    sheet.cell(row=1, column=4).value = 'Reservoir Number'
    sheet.cell(row=1, column=5).value = 'Reservoir Well'
    sheet.cell(row=1, column=6).value = 'Concentration uM'
    sheet.cell(row=1, column=7).value = 'Primer ID'
    sheet.cell(row=1, column=8).value = 'Total Vol Req'
    sheet.cell(row=1, column=9).value = 'Strain'
    sheet.cell(row=1, column=10).value = 'AA'


def save_data_to_csv(file_name_in, file_name_out):
    xlsx = openpyxl.load_workbook(file_name_in)
    sheet = xlsx.active
    data = sheet.rows
    csv = open(file_name_out, "w+")
    for row in data:
        l = list(row)
        for i in range(len(l)):
            if i == len(l) - 1:
                csv.write(str(l[i].value))
            else:
                csv.write(str(l[i].value) + ',')
        csv.write('\n')
    csv.close()
    os.remove(file_name_in)


if __name__ == "__main__":

    path_to_dir = '../Experiments/Sanger_seq_data/Transfer_files_Round3/'
    input_file_names = ['primer_options.xlsx', 'Sequencing Primer Assignment Template.xlsx', 'Round_3_colonies.xlsx']
    input_files = [path_to_dir + name for name in input_file_names]
    plt_num = 1
    output_file = path_to_dir + f'Biomek_transfer_ss_primers_PLT{plt_num}.xlsx'
    output_biomek_file(input_files, output_file)

    # path_to_dir = '../Experiments/Sanger_seq_data/Transfer_files/'
    # input_file_names = ['primer_options.xlsx', 'Sequencing Primer Assignment Template.xlsx', 'R2_PLT1_layout.xlsx']
    # input_files = [path_to_dir + name for name in input_file_names]
    # plt_num = 1
    # output_file = path_to_dir + f'Biomek_transfer_ss_primers_PLT{plt_num}.xlsx'
    # output_biomek_file(input_files, output_file)
    #
    # input_file_names = ['primer_options.xlsx', 'Sequencing Primer Assignment Template.xlsx', 'R2_PLT2_layout.xlsx']
    # input_files = [path_to_dir + name for name in input_file_names]
    # plt_num = 2
    # output_file = path_to_dir + f'Biomek_transfer_ss_primers_PLT{plt_num}.xlsx'
    # output_biomek_file(input_files, output_file)

